import os
import time
import datetime
from tkinter import *
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
#The default theme for custom Tkinter is system defined; themes available are dark-blue, blue and green
ctk.set_default_color_theme("dark-blue")
# EXPORT UPDATES DON'T WORK, 
# OTHER FILES NEED TO BE UPDATED
#Attribute 'config' is not implemented for CTk widgets. For consistency, always use 'configure' instead

# Get the current date
current_date = datetime.date.today().strftime('%d%b%Y')
output_file = f'task_output_{current_date}.xlsx'

# Check if the output Excel file already exists for the current date
if os.path.isfile(output_file):
    workbook = load_workbook(output_file)
else:
    workbook = Workbook()

# Get the active worksheet or create a new one if it doesn't exist
worksheet = workbook.active
if worksheet.title != 'Task Output':
    worksheet.title = 'Task Output'
    worksheet.append(['Task', 'Task Name', 'Duration', 'Client', 'Job'])

# Define the function to start the stopwatch for a task
def start_stopwatch(task, duration, task_name):
    # Create a task_i window
    window = ctk.CTkToplevel(root)
    window.title(task)
    window.geometry('300x300')
    
    # Add a label for the task 
    task_label = ctk.CTkLabel(window, text=task, font=('Arial', 18))
    task_label.pack(pady=10)
    # Add a label for the task name
    input_label = ctk.CTkLabel(window, text=task_name.get(), font=('Arial', 12))
    input_label.pack(pady=5)

    # Add an input field for the user to input Job (task description)
    input_label = ctk.CTkLabel(window, text="Job:", font=('Arial', 12))
    input_label.pack(pady=5)
    ut_entry = ctk.CTkEntry(window)
    ut_entry.pack(pady=5)

    # Add an input field for the user to input Client (name)
    input_label = ctk.CTkLabel(window, text="Client:", font=('Arial', 12))
    input_label.pack(pady=5)
    uc_entry = ctk.CTkEntry(window)
    uc_entry.pack(pady=5)
    
    # Add a label for the stopwatch
    stopwatch_label = ctk.CTkLabel(window, text='00:00:00', font=('Arial', 24))
    stopwatch_label.pack()
    
    # Define the function to update the stopwatch label
    def update_stopwatch():
        nonlocal start_time, elapsed_time
        elapsed_time = time.perf_counter() - start_time
        hours= int(elapsed_time//3600)
        minutes = int((elapsed_time // 60) % 60)
        seconds = int(elapsed_time % 60)
        stopwatch_label.configure(text='{:02d}:{:02d}:{:02d}'.format(hours, minutes, seconds))
        if not final_duration:
            window.after(100, update_stopwatch)
    
     # Define the function to stop the stopwatch and display the final duration
    def stop_stopwatch(event):
        nonlocal final_duration
        final_duration = int(elapsed_time)
        hours = final_duration // 3600
        minutes = (final_duration // 60) % 60
        seconds = final_duration % 60
        
        if hours > 0:
            duration_var.set(f'{hours} hours {minutes} minutes {seconds} seconds')
            duration_label.configure(text=f'{hours} hours {minutes} minutes {seconds} seconds')
            duration_str = f"{hours} hours {minutes} minutes {seconds} seconds"
        elif minutes > 0:
            duration_var.set(f'{minutes} minutes {seconds} seconds')
            duration_label.configure(text=f'{minutes} minutes {seconds} seconds')
            duration_str = f"{minutes} minutes {seconds} seconds"
        else:
            duration_var.set(f'{seconds} seconds')
            duration_label.configure(text=f'{seconds} seconds')
            duration_str = f"{seconds} seconds"
        
        print(f"{task} took {duration_str} in total. Client: {uc_entry.get()}. Job: {ut_entry.get()}")
        # Append the data to the worksheet
        worksheet.append([task, task_name.get(), duration_str, uc_entry.get(), ut_entry.get()])
        workbook.save(output_file)
        window.bind('<Return>', exit_stopwatch)

    def exit_stopwatch(event):
        window.destroy()
        
    # Start the stopwatch
    start_time = time.perf_counter()
    elapsed_time = 0
    final_duration = None
    update_stopwatch()

     # Add a label for the final duration
    duration_label = ctk.CTkLabel(window, text='Hit <Return> to STOP and see the elapsed time.')
    duration_label.pack(pady=10)

    # Bind the <Return> key to stop the stopwatch and display the final duration
    window.bind('<Return>', stop_stopwatch)
    
    # Wait for the user to close the window
    window.mainloop()

# Create the main window
root = ctk.CTk()
root.title('ToDoTi')
root.geometry('500x400')

# Add a label for the to-do list
label = ctk.CTkLabel(root, text='To-Do List', font=('Arial', 18))
label.pack(pady=10)
# Define the tasks and their corresponding duration and input task variables
task_durations = {'Task 1': ctk.StringVar(), 'Task 2': ctk.StringVar(), 'Task 3': ctk.StringVar(), 'Task 4': ctk.StringVar(), 'Task 5': ctk.StringVar()}
task_names = {'Task 1': ctk.StringVar(), 'Task 2': ctk.StringVar(), 'Task 3': ctk.StringVar(), 'Task 4': ctk.StringVar(), Ttask 5': ctk.StringVar()}

# Loop through the tasks and add a button and input field for each task
for task, duration_var in task_durations.items():
    # Create a frame to hold the task and input field
    frame = ctk.CTkFrame(root)
    frame.pack(pady=5)

    # Add the task label
    task_label = ctk.CTkLabel(frame, text=task, font=('Arial', 14))
    task_label.pack(side='left')

    # Add the task name field
    task_name = ctk.CTkEntry(frame, textvariable=task_names[task])
    task_name.pack(side='left', padx=5)

    # Add the button to start the stopwatch
    button = ctk.CTkButton(frame, text='Start', font=('Arial', 12),
                       command=lambda t=task, d=duration_var, f=task_name: start_stopwatch(t, d, f))
    button.pack(side='left')

# Start the main event loop
root.mainloop()
